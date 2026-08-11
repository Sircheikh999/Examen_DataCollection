```python
# ============================================================
# APPLICATION STREAMLIT - EXAMEN DATA COLLECTION
# ============================================================
# Fonctionnalités :
# 1. Scraping Selenium de Books to Scrape et Gaaraas
# 2. Téléchargement des données brutes Web Scraper
# 3. Nettoyage et visualisation des données Selenium
# 4. Accès aux formulaires KoboToolbox et Google Forms
# 5. Stockage SQL des données collectées
# ============================================================

import os
import re
import io
import sqlite3
from pathlib import Path

import pandas as pd
import streamlit as st


# ============================================================
# CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Data Collection Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

BASE_DIR = Path(__file__).resolve().parent

COLAB_DIR = BASE_DIR / "Colab"
WEB_SCRAPER_DIR = BASE_DIR / "Web_Scraper"
FORM_DIR = BASE_DIR / "Formulaire"

DB_PATH = BASE_DIR / "data_collection.db"


# ============================================================
# STYLE
# ============================================================

st.markdown(
    """
    <style>
        .main-title {
            text-align: center;
            font-size: 42px;
            font-weight: bold;
            margin-bottom: 10px;
        }

        .subtitle {
            text-align: center;
            font-size: 18px;
            margin-bottom: 30px;
        }

        .card {
            padding: 20px;
            border-radius: 12px;
            border: 1px solid #ddd;
            margin-bottom: 15px;
        }

        .stButton > button {
            width: 100%;
        }
    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# TITRE
# ============================================================

st.markdown(
    '<div class="main-title">📊 DATA COLLECTION APPLICATION</div>',
    unsafe_allow_html=True
)

st.markdown(
    """
    <div class="subtitle">
    Application de collecte, nettoyage, stockage et visualisation
    des données issues de plusieurs sources.
    </div>
    """,
    unsafe_allow_html=True
)


# ============================================================
# BASE DE DONNÉES SQL
# ============================================================

def get_db_connection():
    """Connexion à la base SQLite."""
    connection = sqlite3.connect(DB_PATH)
    return connection


def dataframe_to_sql(df, table_name):
    """
    Stocke un DataFrame dans une table SQL.
    La table est remplacée à chaque nouvelle collecte.
    """
    if df is None or df.empty:
        return False

    try:
        connection = get_db_connection()

        df.to_sql(
            table_name,
            connection,
            if_exists="replace",
            index=False
        )

        connection.close()

        return True

    except Exception as e:
        st.error(f"Erreur SQL : {e}")
        return False


def get_sql_tables():
    """Retourne la liste des tables de la base."""
    try:
        connection = get_db_connection()

        tables = pd.read_sql_query(
            """
            SELECT name
            FROM sqlite_master
            WHERE type='table'
            ORDER BY name
            """,
            connection
        )

        connection.close()

        return tables["name"].tolist()

    except Exception:
        return []


# ============================================================
# OUTILS DE NETTOYAGE
# ============================================================

def clean_books_data(df):
    """Nettoyage des données Books to Scrape."""

    df = df.copy()

    # Suppression des espaces
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip()

    # Prix numérique
    if "price" in df.columns:
        df["price_numeric"] = (
            df["price"]
            .astype(str)
            .str.replace("£", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.strip()
        )

        df["price_numeric"] = pd.to_numeric(
            df["price_numeric"],
            errors="coerce"
        )

    # Note en nombre
    if "star_rating" in df.columns:
        df["rating"] = (
            df["star_rating"]
            .astype(str)
            .str.extract(
                r"(One|Two|Three|Four|Five)",
                expand=False
            )
        )

        rating_mapping = {
            "One": 1,
            "Two": 2,
            "Three": 3,
            "Four": 4,
            "Five": 5
        }

        df["rating"] = df["rating"].map(rating_mapping)

    # Avis
    if "reviews" in df.columns:
        df["reviews"] = pd.to_numeric(
            df["reviews"],
            errors="coerce"
        ).fillna(0).astype(int)

    # Page
    if "page" in df.columns:
        df["page"] = pd.to_numeric(
            df["page"],
            errors="coerce"
        )

    # Suppression des doublons
    df = df.drop_duplicates()

    return df


def clean_gaaraas_data(df):
    """Nettoyage des données Gaaraas."""

    df = df.copy()

    # Nettoyage des chaînes
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].astype(str).str.strip()

    # Année
    if "annee" in df.columns:
        df["annee"] = pd.to_numeric(
            df["annee"],
            errors="coerce"
        )

    # Prix en FCFA
    if "prix" in df.columns:
        df["prix_numeric"] = (
            df["prix"]
            .astype(str)
            .str.replace("CFA", "", regex=False)
            .str.replace("FCFA", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.strip()
        )

        df["prix_numeric"] = pd.to_numeric(
            df["prix_numeric"],
            errors="coerce"
        )

    # Kilométrage
    if "kilometrage" in df.columns:
        df["kilometrage_numeric"] = (
            df["kilometrage"]
            .astype(str)
            .str.replace("km", "", case=False, regex=False)
            .str.replace("Km", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(",", "", regex=False)
            .str.strip()
        )

        df["kilometrage_numeric"] = pd.to_numeric(
            df["kilometrage_numeric"],
            errors="coerce"
        )

    # Page
    if "page" in df.columns:
        df["page"] = pd.to_numeric(
            df["page"],
            errors="coerce"
        )

    # Suppression des doublons
    df = df.drop_duplicates()

    return df


# ============================================================
# SELENIUM
# ============================================================

def create_driver():
    """
    Création d'un navigateur Chrome Selenium en mode headless.

    Selenium Manager est utilisé afin d'éviter de gérer
    manuellement le ChromeDriver.
    """

    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options

        options = Options()

        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")

        driver = webdriver.Chrome(options=options)

        return driver

    except Exception as e:
        raise RuntimeError(
            "Impossible de lancer Chrome/Selenium. "
            "Vérifiez l'installation de Chrome/Chromium et Selenium. "
            f"Détail : {e}"
        )


# ============================================================
# SCRAPING BOOKS TO SCRAPE
# ============================================================

def scrape_books(start_page=1, end_page=50):

    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    driver = create_driver()

    data = []

    try:

        progress = st.progress(0)

        total_pages = end_page - start_page + 1

        for counter, page in enumerate(
            range(start_page, end_page + 1),
            start=1
        ):

            url = (
                f"https://books.toscrape.com/"
                f"catalogue/page-{page}.html"
            )

            driver.get(url)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "article.product_pod h3 a")
                    )
                )
            except Exception:
                pass

            containers = driver.find_elements(
                By.CSS_SELECTOR,
                "article.product_pod h3 a"
            )

            number_of_products = len(containers)

            product_urls = [
                container.get_attribute("href")
                for container in containers
            ]

            for product_url in product_urls:

                try:

                    driver.get(product_url)

                    item = {
                        "page": page,
                        "number_of_products": number_of_products,
                        "title": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.product_main h1"
                        ).text,
                        "price": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.product_main p.price_color"
                        ).text,
                        "availability": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.product_main p.instock.availability"
                        ).text,
                        "star_rating": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.product_main p.star-rating"
                        ).get_attribute("class"),
                        "reviews": driver.find_element(
                            By.CSS_SELECTOR,
                            "table.table-striped tr:nth-child(7) td"
                        ).text,
                        "description": driver.find_element(
                            By.CSS_SELECTOR,
                            "#product_description + p"
                        ).text,
                        "product_type": driver.find_element(
                            By.CSS_SELECTOR,
                            "ul.breadcrumb li:nth-child(3) a"
                        ).text,
                        "tax": driver.find_element(
                            By.CSS_SELECTOR,
                            "table.table-striped tr:nth-child(5) td"
                        ).text
                    }

                    data.append(item)

                except Exception:
                    continue

            progress.progress(counter / total_pages)

        progress.empty()

    finally:
        driver.quit()

    return pd.DataFrame(data)


# ============================================================
# SCRAPING GAARAAS
# ============================================================

def scrape_gaaraas(start_page=1, end_page=13):

    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    driver = create_driver()

    data = []

    try:

        progress = st.progress(0)

        total_pages = end_page - start_page + 1

        for counter, page in enumerate(
            range(start_page, end_page + 1),
            start=1
        ):

            url = (
                "https://www.gaaraas.com/fr/"
                f"users/dakar-auto?page={page}"
            )

            driver.get(url)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "a.common-ad-card")
                    )
                )
            except Exception:
                pass

            containers = driver.find_elements(
                By.CSS_SELECTOR,
                "a.common-ad-card"
            )

            number_of_ads = len(containers)

            ad_urls = [
                container.get_attribute("href")
                for container in containers
            ]

            for ad_url in ad_urls:

                try:

                    driver.get(ad_url)

                    item = {
                        "page": page,
                        "number_of_ads": number_of_ads,
                        "marque_modele": driver.find_element(
                            By.CSS_SELECTOR,
                            ".ad-title-block h2"
                        ).text,
                        "annee": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.prop:nth-of-type(4) span:nth-of-type(2)"
                        ).text,
                        "prix": driver.find_element(
                            By.CSS_SELECTOR,
                            ".back-wrapper .ad-price span.price-wrap"
                        ).text,
                        "kilometrage": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.prop:nth-of-type(3) span:nth-of-type(2)"
                        ).text,
                        "type_boite_de_vitesse": driver.find_element(
                            By.CSS_SELECTOR,
                            "div.prop:nth-of-type(2) span:nth-of-type(2)"
                        ).text,
                        "region_de_vente": driver.find_element(
                            By.CSS_SELECTOR,
                            ".ad-title a span"
                        ).text,
                        "url": ad_url
                    }

                    data.append(item)

                except Exception:
                    continue

            progress.progress(counter / total_pages)

        progress.empty()

    finally:
        driver.quit()

    return pd.DataFrame(data)


# ============================================================
# WEB SCRAPER
# ============================================================

def find_csv_files():

    if not WEB_SCRAPER_DIR.exists():
        return []

    return sorted(
        WEB_SCRAPER_DIR.glob("*.csv")
    )


def load_csv(file_path):

    try:
        return pd.read_csv(
            file_path,
            encoding="utf-8-sig",
            sep=",",
            on_bad_lines="skip"
        )

    except Exception:

        return pd.read_csv(
            file_path,
            encoding="latin1",
            sep=",",
            on_bad_lines="skip"
        )


# ============================================================
# FORMULAIRES
# ============================================================

def extract_urls_from_excel(file_path):
    """
    Recherche des liens dans les fichiers Excel.
    Les liens peuvent être présents :
    - sous forme d'hyperliens Excel
    - directement sous forme de texte.
    """

    urls = []

    try:

        from openpyxl import load_workbook

        workbook = load_workbook(
            file_path,
            data_only=False
        )

        for worksheet in workbook.worksheets:

            for row in worksheet.iter_rows():

                for cell in row:

                    # Hyperlink Excel
                    if cell.hyperlink:
                        target = cell.hyperlink.target

                        if target and target not in urls:
                            urls.append(target)

                    # URL écrite dans la cellule
                    if isinstance(cell.value, str):

                        matches = re.findall(
                            r"https?://[^\s\"'<>]+",
                            cell.value
                        )

                        for url in matches:

                            if url not in urls:
                                urls.append(url)

        workbook.close()

    except Exception:
        pass

    return urls


def get_form_links():

    links = {
        "Google Forms": [],
        "KoboToolbox": []
    }

    google_file = FORM_DIR / "formulaire_forms.xlsx"
    kobo_file = FORM_DIR / "formulaire_kobotoolbox.xlsx"

    if google_file.exists():
        links["Google Forms"] = extract_urls_from_excel(
            google_file
        )

    if kobo_file.exists():
        links["KoboToolbox"] = extract_urls_from_excel(
            kobo_file
        )

    return links


# ============================================================
# SIDEBAR
# ============================================================

st.sidebar.title("🧭 Navigation")

page = st.sidebar.radio(
    "Choisir une section",
    [
        "🏠 Accueil",
        "🕷️ Scraping Selenium",
        "📥 Données Web Scraper",
        "📊 Dashboard",
        "📝 Évaluation",
        "🗄️ Base SQL"
    ]
)


# ============================================================
# ACCUEIL
# ============================================================

if page == "🏠 Accueil":

    st.header("Bienvenue dans l'application")

    st.markdown(
        """
        Cette application permet de centraliser les différentes
        étapes du projet Data Collection :

        ### 🕷️ 1. Scraping Selenium
        - Books to Scrape
        - Gaaraas
        - Navigation sur plusieurs pages
        - Extraction des données détaillées

        ### 📥 2. Web Scraper
        Téléchargement des deux fichiers CSV bruts produits
        avec l'outil no-code Web Scraper.

        ### 🧹 3. Nettoyage
        Les données Selenium sont nettoyées avant leur
        exploitation dans le dashboard.

        ### 📊 4. Dashboard
        Visualisation interactive :
        - nombre d'observations
        - prix
        - catégories
        - années
        - kilométrage
        - notes

        ### 📝 5. Évaluation
        Accès direct aux deux formulaires :
        - Google Forms
        - KoboToolbox

        ### 🗄️ 6. Base SQL
        Les données collectées peuvent être enregistrées
        automatiquement dans une base SQLite.
        """
    )

    st.success(
        "Application prête pour le déploiement Streamlit."
    )


# ============================================================
# SCRAPING SELENIUM
# ============================================================

elif page == "🕷️ Scraping Selenium":

    st.header("🕷️ Scraping des données avec Selenium")

    tab1, tab2 = st.tabs(
        [
            "📚 Books to Scrape",
            "🚗 Gaaraas"
        ]
    )

    # --------------------------------------------------------
    # BOOKS
    # --------------------------------------------------------

    with tab1:

        st.subheader("📚 Books to Scrape")

        st.write(
            "Le notebook Selenium du projet parcourt les "
            "50 pages de Books to Scrape."
        )

        col1, col2 = st.columns(2)

        with col1:
            books_start = st.number_input(
                "Page de départ",
                min_value=1,
                max_value=50,
                value=1,
                key="books_start"
            )

        with col2:
            books_end = st.number_input(
                "Page finale",
                min_value=1,
                max_value=50,
                value=50,
                key="books_end"
            )

        if st.button(
            "🚀 Lancer le scraping Books to Scrape",
            key="scrape_books"
        ):

            if books_start > books_end:

                st.error(
                    "La page de départ doit être inférieure "
                    "ou égale à la page finale."
                )

            else:

                with st.spinner(
                    "Scraping Books to Scrape en cours..."
                ):

                    try:

                        books_df = scrape_books(
                            books_start,
                            books_end
                        )

                        books_clean = clean_books_data(
                            books_df
                        )

                        st.session_state[
                            "books_selenium"
                        ] = books_clean

                        dataframe_to_sql(
                            books_clean,
                            "books_selenium"
                        )

                        st.success(
                            f"{len(books_clean)} livres récupérés."
                        )

                    except Exception as e:

                        st.error(
                            f"Erreur pendant le scraping : {e}"
                        )

        if "books_selenium" in st.session_state:

            df = st.session_state["books_selenium"]

            st.metric(
                "Nombre de livres",
                len(df)
            )

            st.dataframe(
                df,
                use_container_width=True
            )

            csv = df.to_csv(index=False).encode("utf-8")

            st.download_button(
                "⬇️ Télécharger les données Selenium",
                csv,
                "books_selenium_clean.csv",
                "text/csv",
                key="download_books"
            )

    # --------------------------------------------------------
    # GAARAAS
    # --------------------------------------------------------

    with tab2:

        st.subheader("🚗 Gaaraas")

        st.write(
            "Le notebook Selenium du projet parcourt "
            "les 13 pages du vendeur Dakar Auto."
        )

        col1, col2 = st.columns(2)

        with col1:

            gaaraas_start = st.number_input(
                "Page de départ",
                min_value=1,
                max_value=13,
                value=1,
                key="gaaraas_start"
            )

        with col2:

            gaaraas_end = st.number_input(
                "Page finale",
                min_value=1,
                max_value=13,
                value=13,
                key="gaaraas_end"
            )

        if st.button(
            "🚀 Lancer le scraping Gaaraas",
            key="scrape_gaaraas"
        ):

            if gaaraas_start > gaaraas_end:

                st.error(
                    "La page de départ doit être inférieure "
                    "ou égale à la page finale."
                )

            else:

                with st.spinner(
                    "Scraping Gaaraas en cours..."
                ):

                    try:

                        gaaraas_df = scrape_gaaraas(
                            gaaraas_start,
                            gaaraas_end
                        )

                        gaaraas_clean = clean_gaaraas_data(
                            gaaraas_df
                        )

                        st.session_state[
                            "gaaraas_selenium"
                        ] = gaaraas_clean

                        dataframe_to_sql(
                            gaaraas_clean,
                            "gaaraas_selenium"
                        )

                        st.success(
                            f"{len(gaaraas_clean)} annonces récupérées."
                        )

                    except Exception as e:

                        st.error(
                            f"Erreur pendant le scraping : {e}"
                        )

        if "gaaraas_selenium" in st.session_state:

            df = st.session_state["gaaraas_selenium"]

            st.metric(
                "Nombre d'annonces",
                len(df)
            )

            st.dataframe(
                df,
                use_container_width=True
            )

            csv = df.to_csv(index=False).encode("utf-8")

            st.download_button(
                "⬇️ Télécharger les données Selenium",
                csv,
                "gaaraas_selenium_clean.csv",
                "text/csv",
                key="download_gaaraas"
            )


# ============================================================
# WEB SCRAPER
# ============================================================

elif page == "📥 Données Web Scraper":

    st.header("📥 Données brutes Web Scraper")

    st.write(
        "Les fichiers CSV présents dans le dossier "
        "`Web_Scraper` sont disponibles ici."
    )

    csv_files = find_csv_files()

    if not csv_files:

        st.warning(
            "Aucun fichier CSV trouvé dans Web_Scraper."
        )

    else:

        for file_path in csv_files:

            with st.expander(
                f"📄 {file_path.name}"
            ):

                try:

                    df = load_csv(file_path)

                    col1, col2 = st.columns(2)

                    with col1:
                        st.metric(
                            "Lignes",
                            df.shape[0]
                        )

                    with col2:
                        st.metric(
                            "Colonnes",
                            df.shape[1]
                        )

                    st.dataframe(
                        df.head(100),
                        use_container_width=True
                    )

                    csv_data = df.to_csv(
                        index=False
                    ).encode("utf-8")

                    st.download_button(
                        "⬇️ Télécharger le CSV brut",
                        data=csv_data,
                        file_name=file_path.name,
                        mime="text/csv",
                        key=f"download_{file_path.name}"
                    )

                    # Stockage SQL de la source Web Scraper
                    if "Books" in file_path.name:

                        dataframe_to_sql(
                            df,
                            "books_webscraper"
                        )

                    elif "Gaaraas" in file_path.name:

                        dataframe_to_sql(
                            df,
                            "gaaraas_webscraper"
                        )

                except Exception as e:

                    st.error(
                        f"Impossible de lire le fichier : {e}"
                    )


# ============================================================
# DASHBOARD
# ============================================================

elif page == "📊 Dashboard":

    st.header("📊 Dashboard des données nettoyées")

    st.info(
        "Le dashboard utilise les données obtenues par "
        "le scraping Selenium et nettoyées par l'application."
    )

    # ========================================================
    # BOOKS DASHBOARD
    # ========================================================

    if "books_selenium" in st.session_state:

        st.subheader("📚 Books to Scrape")

        books = st.session_state[
            "books_selenium"
        ].copy()

        c1, c2, c3, c4 = st.columns(4)

        with c1:
            st.metric(
                "Livres",
                len(books)
            )

        with c2:
            if "price_numeric" in books:
                st.metric(
                    "Prix moyen",
                    f"£{books['price_numeric'].mean():.2f}"
                )

        with c3:
            if "rating" in books:
                st.metric(
                    "Note moyenne",
                    f"{books['rating'].mean():.2f}/5"
                )

        with c4:
            if "product_type" in books:
                st.metric(
                    "Catégories",
                    books["product_type"].nunique()
                )

        col1, col2 = st.columns(2)

        with col1:

            if "product_type" in books.columns:

                category_count = (
                    books["product_type"]
                    .value_counts()
                    .head(10)
                )

                st.write("### Top catégories")

                st.bar_chart(
                    category_count
                )

        with col2:

            if "rating" in books.columns:

                rating_count = (
                    books["rating"]
                    .value_counts()
                    .sort_index()
                )

                st.write("### Répartition des notes")

                st.bar_chart(
                    rating_count
                )

        st.write("### Données nettoyées")

        st.dataframe(
            books,
            use_container_width=True
        )

    else:

        st.warning(
            "Aucune donnée Books to Scrape disponible. "
            "Lancez le scraping dans la section Selenium."
        )

    st.divider()

    # ========================================================
    # GAARAAS DASHBOARD
    # ========================================================

    if "gaaraas_selenium" in st.session_state:

        st.subheader("🚗 Gaaraas")

        gaaraas = st.session_state[
            "gaaraas_selenium"
        ].copy()

        c1, c2, c3, c4 = st.columns(4)

        with c1:
            st.metric(
                "Annonces",
                len(gaaraas)
            )

        with c2:

            if "prix_numeric" in gaaraas:

                st.metric(
                    "Prix moyen",
                    f"{gaaraas['prix_numeric'].mean():,.0f} CFA"
                )

        with c3:

            if "annee" in gaaraas:

                st.metric(
                    "Année moyenne",
                    f"{gaaraas['annee'].mean():.0f}"
                )

        with c4:

            if "kilometrage_numeric" in gaaraas:

                st.metric(
                    "Kilométrage moyen",
                    f"{gaaraas['kilometrage_numeric'].mean():,.0f} km"
                )

        col1, col2 = st.columns(2)

        with col1:

            if "marque_modele" in gaaraas:

                brands = (
                    gaaraas["marque_modele"]
                    .value_counts()
                    .head(10)
                )

                st.write(
                    "### Marques / modèles les plus représentés"
                )

                st.bar_chart(
                    brands
                )

        with col2:

            if "annee" in gaaraas:

                years = (
                    gaaraas["annee"]
                    .value_counts()
                    .sort_index()
                )

                st.write(
                    "### Répartition par année"
                )

                st.bar_chart(
                    years
                )

        st.write("### Données nettoyées")

        st.dataframe(
            gaaraas,
            use_container_width=True
        )

    else:

        st.warning(
            "Aucune donnée Gaaraas disponible. "
            "Lancez le scraping dans la section Selenium."
        )


# ============================================================
# ÉVALUATION
# ============================================================

elif menu == "Formulaires d'évaluation":

    st.title("📝 Formulaires d'évaluation")

    st.write(
        "Les deux formulaires d'évaluation sont accessibles "
        "directement depuis l'application."
    )

    st.subheader("📋 Google Forms")

    st.link_button(
        "📝 Ouvrir le formulaire Google Forms",
        "https://docs.google.com/forms/d/e/1FAIpQLSc7F8m3eBJkCqpOUa4pTQX0zyIov_4LWXRYOV3XKbmi0vJJoQ/viewform?usp=publish-editor",
        use_container_width=True,
    )

    st.divider()

    st.subheader("📋 KoboToolbox")

    st.link_button(
        "📝 Ouvrir le formulaire KoboToolbox",
        "https://ee.kobotoolbox.org/i/1zbGqqaq",
        use_container_width=True,
    )


# ============================================================
# BASE SQL
# ============================================================

elif page == "🗄️ Base SQL":

    st.header("🗄️ Base de données SQL")

    st.write(
        """
        L'application utilise une base SQLite nommée
        `data_collection.db`.

        Les différentes sources sont stockées dans des
        tables séparées afin de conserver leur structure
        respective.
        """
    )

    tables = get_sql_tables()

    if not tables:

        st.info(
            "La base est actuellement vide. "
            "Lancez un scraping ou ouvrez les données "
            "Web Scraper pour alimenter la base."
        )

    else:

        st.success(
            f"{len(tables)} table(s) disponible(s)."
        )

        for table in tables:

            with st.expander(
                f"📁 Table : {table}"
            ):

                try:

                    connection = get_db_connection()

                    df = pd.read_sql_query(
                        f'SELECT * FROM "{table}"',
                        connection
                    )

                    connection.close()

                    st.write(
                        f"Dimensions : "
                        f"{df.shape[0]} lignes × "
                        f"{df.shape[1]} colonnes"
                    )

                    st.dataframe(
                        df.head(100),
                        use_container_width=True
                    )

                    csv = df.to_csv(
                        index=False
                    ).encode("utf-8")

                    st.download_button(
                        "⬇️ Télécharger cette table",
                        csv,
                        f"{table}.csv",
                        "text/csv",
                        key=f"sql_download_{table}"
                    )

                except Exception as e:

                    st.error(
                        f"Erreur de lecture SQL : {e}"
                    )

    st.divider()

    st.subheader("📌 Schéma de stockage")

    schema = pd.DataFrame(
        {
            "Source": [
                "Selenium - Books to Scrape",
                "Selenium - Gaaraas",
                "Web Scraper - Books",
                "Web Scraper - Gaaraas"
            ],
            "Table SQL": [
                "books_selenium",
                "gaaraas_selenium",
                "books_webscraper",
                "gaaraas_webscraper"
            ]
        }
    )

    st.dataframe(
        schema,
        use_container_width=True,
        hide_index=True
    )


# ============================================================
# FOOTER
# ============================================================

st.sidebar.divider()

st.sidebar.caption(
    "Projet Examen Data Collection"
)

st.sidebar.caption(
    "Streamlit • Selenium • Pandas • SQLite"
)
```
